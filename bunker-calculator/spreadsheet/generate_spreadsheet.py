#!/usr/bin/env python3
"""Generate bunker fuel calculator Excel workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "bunker-fuel-calculator.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_instructions(ws):
    ws.title = "Instructions"
    ws.column_dimensions["A"].width = 90
    lines = [
        "BUNKER FUEL CALCULATOR — Instructions",
        "",
        "1. Go to the 'Tank_Table' sheet and replace the sample values with your engine-room sounding table.",
        "   • Row 1 (B1:F1): trim values in metres (by stern = positive).",
        "   • Column A (A2:A12): sounding values in metres.",
        "   • Grid cells: volume in m³ at each sounding/trim intersection.",
        "   • Cell B14: list correction in m³ per degree (0 if not used).",
        "",
        "2. Go to 'Calculator' sheet.",
        "   • Enter ship trim (m), list (°), and density @ 15°C (kg/m³) in the yellow cells.",
        "   • Enter each tank sounding (m) and observed temperature (°C).",
        "   • Observed volume, VCF, standard volume @ 15°C, and mass (MT) calculate automatically.",
        "",
        "3. Formulas used (same as Bunker Master / ASTM 54B):",
        "   • Volume: bilinear interpolation from Tank_Table using sounding + trim, plus list correction.",
        "   • VCF = EXP(-613.9723 / density² × (temp - 15))",
        "   • Mass (MT) = standard volume × density @ 15°C / 1000",
        "",
        "4. For additional tanks with different capacity tables, duplicate this file per tank",
        "   or add extra table sheets and adjust the Calculator formulas.",
        "",
        "Web app version: ../web/index.html (open in browser, import CSV tank table).",
    ]
    for i, line in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=line)
        if i == 1:
            ws.cell(row=i, column=1).font = Font(bold=True, size=14)


def build_tank_table(ws):
    ws.title = "Tank_Table"
    trims = [0.0, 1.0, 2.0, 3.0, 4.0]
    ws["A1"] = "Sounding \\ Trim (m)"
    for j, t in enumerate(trims, 2):
        ws.cell(row=1, column=j, value=t)
    style_header(ws, 1, 1, len(trims) + 1)

    soundings = [6.0, 6.5, 7.0, 7.5, 8.0, 8.5, 9.0, 9.5, 10.0]
    sample_vols = [
        [380.5, 378.2, 376.0, 373.8, 371.5],
        [395.2, 392.9, 390.7, 388.4, 386.2],
        [410.0, 407.7, 405.5, 403.2, 401.0],
        [424.8, 422.5, 420.3, 418.0, 415.8],
        [439.5, 437.2, 435.0, 432.7, 430.5],
        [454.3, 452.0, 449.8, 447.5, 445.3],
        [469.0, 466.7, 464.5, 462.2, 460.0],
        [483.8, 481.5, 479.3, 477.0, 474.8],
        [498.5, 496.2, 494.0, 491.7, 489.5],
    ]
    for i, (s, row) in enumerate(zip(soundings, sample_vols), 2):
        ws.cell(row=i, column=1, value=s)
        for j, v in enumerate(row, 2):
            ws.cell(row=i, column=j, value=v)

    ws["A14"] = "List correction (m³/deg)"
    ws["B14"] = 0.15
    ws["A16"] = "Tank name"
    ws["B16"] = "No.1 HFO Service"

    ws.column_dimensions["A"].width = 18
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 12


def build_calculator(wb, ws):
    ws.title = "Calculator"
    yellow = PatternFill("solid", fgColor="FFF2CC")

    ws["A1"] = "BUNKER FUEL CALCULATOR"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Trim (m, by stern +)"
    ws["B3"] = 2.0
    ws["A4"] = "List (°, stbd +)"
    ws["B4"] = 0
    ws["A5"] = "Density @ 15°C (kg/m³)"
    ws["B5"] = 991
    for r in (3, 4, 5):
        ws.cell(row=r, column=2).fill = yellow

    headers = [
        "Tank",
        "Sounding (m)",
        "Temp (°C)",
        "Obs. vol (m³)",
        "VCF",
        "Std vol @ 15°C (m³)",
        "Mass (MT)",
    ]
    start_row = 7
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, start_row, 1, len(headers))

    tanks = [
        ("No.1 HFO Service", 8.5, 45),
        ("No.2 HFO Settling", 7.2, 42),
        ("No.3 HFO Storage", 9.0, 38),
    ]

    # Named ranges for tank table dimensions
    # Soundings: Tank_Table!$A$2:$A$10
    # Trims: Tank_Table!$B$1:$F$1
    # Volumes: Tank_Table!$B$2:$F$10

    for i, (name, sounding, temp) in enumerate(tanks, start_row + 1):
        r = i
        ws.cell(row=r, column=1, value=name).fill = yellow
        ws.cell(row=r, column=2, value=sounding).fill = yellow
        ws.cell(row=r, column=3, value=temp).fill = yellow

        s_cell = f"$B{r}"
        t_cell = f"$C{r}"

        # Bilinear volume lookup via helper approach using FORECAST.LINEAR
        # Step 1: find lower sounding row
        # MATCH(sounding, soundings, 1) gives index of largest value <= sounding
        vol_formula = (
            f'=LET('
            f's,{s_cell},'
            f'trim,$B$3,'
            f'list,$B$4,'
            f'listCorr,Tank_Table!$B$14,'
            f'sIdx,MATCH(s,Tank_Table!$A$2:$A$10,1),'
            f'tIdx,MATCH(trim,Tank_Table!$B$1:$F$1,1),'
            f's0,INDEX(Tank_Table!$A$2:$A$10,sIdx),'
            f's1,INDEX(Tank_Table!$A$2:$A$10,sIdx+1),'
            f't0,INDEX(Tank_Table!$B$1:$F$1,tIdx),'
            f't1,INDEX(Tank_Table!$B$1:$F$1,tIdx+1),'
            f'v00,INDEX(Tank_Table!$B$2:$F$10,sIdx,tIdx),'
            f'v01,INDEX(Tank_Table!$B$2:$F$10,sIdx,tIdx+1),'
            f'v10,INDEX(Tank_Table!$B$2:$F$10,sIdx+1,tIdx),'
            f'v11,INDEX(Tank_Table!$B$2:$F$10,sIdx+1,tIdx+1),'
            f'vS0,IF(s1=s0,v00,v00+(s-s0)/(s1-s0)*(v10-v00)),'
            f'vS1,IF(s1=s0,v01,v01+(s-s0)/(s1-s0)*(v11-v01)),'
            f'obs,IF(t1=t0,vS0,vS0+(trim-t0)/(t1-t0)*(vS1-vS0)),'
            f'obs+list*listCorr)'
        )
        ws.cell(row=r, column=4, value=vol_formula)

        vcf_formula = f"=EXP(-613.9723/($B$5^2)*({t_cell}-15))"
        ws.cell(row=r, column=5, value=vcf_formula)

        std_formula = f"=D{r}*E{r}"
        ws.cell(row=r, column=6, value=std_formula)

        mt_formula = f"=F{r}*$B$5/1000"
        ws.cell(row=r, column=7, value=mt_formula)

        for c in range(4, 8):
            ws.cell(row=r, column=c).number_format = "0.000"

    total_row = start_row + len(tanks) + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=f"=SUM(D{start_row+1}:D{total_row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F{start_row+1}:F{total_row-1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G{start_row+1}:G{total_row-1})")
    ws.cell(row=total_row, column=7).font = Font(bold=True)

    ws.column_dimensions["A"].width = 22
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 16

    ws["A20"] = "Add more tank rows above TOTAL and copy formulas from row 8."
    ws["A20"].font = Font(italic=True, color="666666")


def main():
    wb = Workbook()
    build_instructions(wb.active)
    build_tank_table(wb.create_sheet())
    build_calculator(wb, wb.create_sheet())
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
